# ----------------------------------------------------------------------------------------------------------------------
# IMPORT GENERAL PACKAGES
# ----------------------------------------------------------------------------------------------------------------------

import os
import sys
import pandas as pd
import yaml
from openpyxl import load_workbook
import datetime as dt

# ----------------------------------------------------------------------------------------------------------------------
# IMPORT PCSE PACKAGES
# ----------------------------------------------------------------------------------------------------------------------

import openpyxl
import sqlalchemy
import traitlets_pcse
import requests
import xlrd

# del sys.path[0:20]
# path_to_the_model = os.path.abspath(os.path.join(os.getcwd(), './# pyWOFOST/pcse-master'))
# sys.path.append(path_to_the_model)

import pcse
from pcse.fileinput import ExcelWeatherDataProvider
from pcse.db import NASAPowerWeatherDataProvider
from pcse.db import AgERA5WeatherDataProvider

# ----------------------------------------------------------------------------------------------------------------------
# SPECIFY DIRECTORY
# ----------------------------------------------------------------------------------------------------------------------

input_dir = r'C:\# Jvasco\Working Papers\# Global Yield Gap Analysis\GYGA_Egypt_Wheat\data\wofost-inputs'

# ----------------------------------------------------------------------------------------------------------------------
# DEFINE SIMULATIONS
# ----------------------------------------------------------------------------------------------------------------------

sites = ['NileDelta', 'UpperEgypt']
coord = [(31.00, 31.00), (32.639637, 25.687243)]   # lon, lat
locat = zip(sites, coord)

# ----------------------------------------------------------------------------------------------------------------------
# GET NASA WEATHER
# ----------------------------------------------------------------------------------------------------------------------

for site, gps in locat:
    print("Running for . . . {site} - {gps}".format(site=site, gps=gps))
    weather_era5 = AgERA5WeatherDataProvider(latitude=gps[1], longitude=gps[0], start_date=dt.date(1980, 1, 1), enddate=dt.date(2020, 6, 1))
    era5_export = pd.DataFrame(weather_era5.export()).set_index('DAY')
    era5_export.to_excel(os.path.join(input_dir, './weather_agera5_{site}.xlsx'.format(site=site)))
    # weather_nasa = NASAPowerWeatherDataProvider(latitude=gps[1], longitude=gps[0], force_update=True)
    # angstrom = pd.DataFrame(data={'Angst_A': weather_nasa.angstA, 'Angst_B': weather_nasa.angstB}, index=[0])
    # angstrom.to_excel(os.path.join(input_dir, './angstrom_coef_{site}.xlsx'.format(site=site)))
    # nasa_export = pd.DataFrame(weather_nasa.export()).set_index('DAY')
    # nasa_export.to_excel(os.path.join(input_dir, './weather_nasa_{site}.xlsx'.format(site=site)))

# ----------------------------------------------------------------------------------------------------------------------
# NASA WEATHER TEMPLATE
# ----------------------------------------------------------------------------------------------------------------------

fn = os.path.join(input_dir, '.\weather_template.xlsx')
template = pd.read_excel(fn, header=None, sheet_name='ObservedWeather')
for site, gps in locat:
    # station = pd.read_excel(os.path.join(input_dir, '.\weather_nasa_{site}.xlsx'.format(site=site)), sheet_name='Sheet1')
    station = pd.read_excel(os.path.join(input_dir, '.\weather_agera5_{site}.xlsx'.format(site=site)), sheet_name='Sheet1')
    r = pd.date_range(start=station['DAY'].min(), end=station['DAY'].max())
    station = station.set_index('DAY').reindex(r).rename_axis('DAY').reset_index()
    station = station[['DAY', 'IRRAD', 'TMIN', 'TMAX', 'VAP', 'WIND', 'RAIN']]
    station['DAY'] = station['DAY'].dt.date
    station['IRRAD'] = station['IRRAD'] / 1000    # J to kJ
    station['RAIN'] = station['RAIN'] * 10        # cm to mm
    station['VAP'] = station['VAP'] / 10          # hPa to kPa
    station['SNOWDEPTH'] = 0
    angstrom = pd.read_excel(os.path.join(input_dir, r'.\angstrom_coef_{site}.xlsx'.format(site=site)), sheet_name='Sheet1')
    writer = pd.ExcelWriter(fn, engine='openpyxl')
    book = load_workbook(fn)
    writer.book = book
    template.to_excel(writer, sheet_name=site, header=None, index=False)
    station.to_excel(writer, sheet_name=site, header=False, index=False, startcol=0, startrow=12)
    sheetname = book[site]
    sheetname.cell(row=2, column=2).value = 'Egypt'
    sheetname.cell(row=3, column=2).value = site
    sheetname.cell(row=4, column=2).value = 'Processed by JV Silva'
    # sheetname.cell(row=5, column=2).value = 'NASA Power'
    sheetname.cell(row=5, column=2).value = 'AgERA 5'
    sheetname.cell(row=6, column=2).value = 'Joao Vasco Silva, WUR'
    sheetname.cell(row=9, column=1).value = gps[0]
    sheetname.cell(row=9, column=2).value = gps[1]
    sheetname.cell(row=9, column=3).value = -99
    sheetname.cell(row=9, column=4).value = angstrom['Angst_A'].unique()[0]
    sheetname.cell(row=9, column=5).value = angstrom['Angst_B'].unique()[0]
    writer.book.remove_sheet(book.get_sheet_by_name('ObservedWeather'))
    # writer.book.save(os.path.join(input_dir, '.\weather_nasa_final_{site}.xlsx'.format(site=site)))
    writer.book.save(os.path.join(input_dir, '.\weather_agera5_final_{site}.xlsx'.format(site=site)))
    writer.book.close()

# ----------------------------------------------------------------------------------------------------------------------
# THE END
# ----------------------------------------------------------------------------------------------------------------------
