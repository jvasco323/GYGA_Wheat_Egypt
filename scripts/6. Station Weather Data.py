# ----------------------------------------------------------------------------------------------------------------------
# IMPORT GENERAL PACKAGES
# ----------------------------------------------------------------------------------------------------------------------

import os
import sys
import numpy as np
import pandas as pd
import yaml
from openpyxl import load_workbook
import datetime as dt

# ----------------------------------------------------------------------------------------------------------------------
# IMPORT PCSE PACKAGES
# ----------------------------------------------------------------------------------------------------------------------

import openpyxl
import sqlalchemy
import traitlets_pcse
import requests
import xlrd

# del sys.path[0:20]
# path_to_the_model = os.path.abspath(os.path.join(os.getcwd(), './# pyWOFOST/pcse-master'))
# sys.path.append(path_to_the_model)

import pcse
from pcse.fileinput import ExcelWeatherDataProvider
from pcse.db import NASAPowerWeatherDataProvider
from pcse.db import AgERA5WeatherDataProvider
from pcse.util import ea_from_tdew

# ----------------------------------------------------------------------------------------------------------------------
# SPECIFY DIRECTORY
# ----------------------------------------------------------------------------------------------------------------------

input_dir = r'D:\# Jvasco\Working Papers\# CIMMYT Database\GYGA Global Yield Gap Atlas\GYGA_Egypt_Wheat\data'

# ----------------------------------------------------------------------------------------------------------------------
# CLEAN NILE DELTA
# ----------------------------------------------------------------------------------------------------------------------

NileDelta = pd.read_excel(os.path.join(input_dir, '.\weather-stations\Climate data-Egypt-Kafr  El-Sheik-Daily-2009-2019.xlsx'), sheet_name='Kafr El sheik')
NileDelta = NileDelta.drop(['Unnamed: 7', 'Unnamed: 8'], axis=1)
NileDelta.columns = ['DAY', 'IRRAD', 'TMIN', 'TMAX', 'VAP', 'WIND', 'RAIN']
NileDelta = NileDelta.iloc[11:].reset_index()
NileDelta['TEMP'] = (NileDelta['TMAX']+NileDelta['TMIN']) / 2
NileDelta['IRRAD'] = NileDelta['IRRAD'] * 1000
NileDelta['RAIN'] = np.where(NileDelta['RAIN'] > 25, 24.5, NileDelta['RAIN'])
NileDelta['VAP'] = NileDelta.apply(lambda df: ea_from_tdew(df['TMIN']), axis=1)
NileDelta['DAY'] = pd.to_datetime(NileDelta['DAY'])
NileDelta['DAY'] = NileDelta['DAY'].dt.date
NileDelta = NileDelta.drop(['index'], axis=1)
NileDelta.to_excel(os.path.join(input_dir, '.\wofost-inputs\weather_actual_NileDelta_unformatted.xlsx'), index=False)

fn = os.path.join(input_dir, '.\wofost-inputs\weather_template.xlsx')
template = pd.read_excel(fn, header=None, sheet_name='ObservedWeather')
r = pd.date_range(start=NileDelta['DAY'].min(), end=NileDelta['DAY'].max())
NileDelta = NileDelta.set_index('DAY').reindex(r).rename_axis('DAY').reset_index()
NileDelta = NileDelta[['DAY', 'IRRAD', 'TMIN', 'TMAX', 'VAP', 'WIND', 'RAIN']]
NileDelta['DAY'] = NileDelta['DAY'].dt.date
NileDelta['IRRAD'] = NileDelta['IRRAD']           # MJ
NileDelta['RAIN'] = NileDelta['RAIN']             # mm
NileDelta['VAP'] = NileDelta['VAP']               # kPa
NileDelta['SNOWDEPTH'] = 0
angstrom = pd.read_excel(os.path.join(input_dir, r'.\wofost-inputs\angstrom_coef_NileDelta.xlsx'), sheet_name='Sheet1')
writer = pd.ExcelWriter(fn, engine='openpyxl')
book = load_workbook(fn)
writer.book = book
template.to_excel(writer, sheet_name='NileDelta', header=None, index=False)
NileDelta.to_excel(writer, sheet_name='NileDelta', header=False, index=False, startcol=0, startrow=12)
sheetname = book['NileDelta']
sheetname.cell(row=2, column=2).value = 'Egypt'
sheetname.cell(row=3, column=2).value = 'NileDelta'
sheetname.cell(row=4, column=2).value = 'Processed by JV Silva'
sheetname.cell(row=5, column=2).value = 'ICARDA'
sheetname.cell(row=6, column=2).value = 'Joao Vasco Silva, WUR'
sheetname.cell(row=9, column=1).value = 31.00
sheetname.cell(row=9, column=2).value = 31.00
sheetname.cell(row=9, column=3).value = -99
sheetname.cell(row=9, column=4).value = angstrom['Angst_A'].unique()[0]
sheetname.cell(row=9, column=5).value = angstrom['Angst_B'].unique()[0]
writer.book.remove_sheet(book.get_sheet_by_name('ObservedWeather'))
writer.book.save(os.path.join(input_dir, '.\wofost-inputs\weather_actual_NileDelta.xlsx'))
writer.book.close()

# ----------------------------------------------------------------------------------------------------------------------
# CLEAN UPPER EGYPT
# ----------------------------------------------------------------------------------------------------------------------

UpperEgypt = pd.read_excel(os.path.join(input_dir, '.\weather-stations\Climate  data-Egypt-Luxor-Daily-2009-2019.xlsx'), sheet_name='Luxor')
UpperEgypt = UpperEgypt.drop(['Unnamed: 7', 'Unnamed: 8'], axis=1)
UpperEgypt.columns = ['DAY', 'IRRAD', 'TMIN', 'TMAX', 'VAP', 'WIND', 'RAIN']
UpperEgypt = UpperEgypt.iloc[11:].reset_index()
UpperEgypt['TEMP'] = (UpperEgypt['TMAX']+UpperEgypt['TMIN']) / 2
UpperEgypt['IRRAD'] = UpperEgypt['IRRAD'] * 1000
UpperEgypt['RAIN'] = np.where(UpperEgypt['RAIN'] == -99, 0, UpperEgypt['RAIN'])
UpperEgypt['VAP'] = UpperEgypt.apply(lambda df: ea_from_tdew(df['TMIN']), axis=1)
UpperEgypt['DAY'] = pd.to_datetime(UpperEgypt['DAY'])
UpperEgypt['DAY'] = UpperEgypt['DAY'].dt.date
UpperEgypt = UpperEgypt.drop(['index'], axis=1)
UpperEgypt.to_excel(os.path.join(input_dir, '.\wofost-inputs\weather_actual_UpperEgypt_unformatted.xlsx'), index=False)

fn = os.path.join(input_dir, '.\wofost-inputs\weather_template.xlsx')
template = pd.read_excel(fn, header=None, sheet_name='ObservedWeather')
r = pd.date_range(start=UpperEgypt['DAY'].min(), end=UpperEgypt['DAY'].max())
UpperEgypt = UpperEgypt.set_index('DAY').reindex(r).rename_axis('DAY').reset_index()
UpperEgypt = UpperEgypt[['DAY', 'IRRAD', 'TMIN', 'TMAX', 'VAP', 'WIND', 'RAIN']]
UpperEgypt['DAY'] = UpperEgypt['DAY'].dt.date
UpperEgypt['IRRAD'] = UpperEgypt['IRRAD']           # MJ
UpperEgypt['RAIN'] = UpperEgypt['RAIN']             # mm
UpperEgypt['VAP'] = UpperEgypt['VAP']               # kPa
UpperEgypt['SNOWDEPTH'] = 0
angstrom = pd.read_excel(os.path.join(input_dir, r'.\wofost-inputs\angstrom_coef_UpperEgypt.xlsx'), sheet_name='Sheet1')
writer = pd.ExcelWriter(fn, engine='openpyxl')
book = load_workbook(fn)
writer.book = book
template.to_excel(writer, sheet_name='UpperEgypt', header=None, index=False)
UpperEgypt.to_excel(writer, sheet_name='UpperEgypt', header=False, index=False, startcol=0, startrow=12)
sheetname = book['UpperEgypt']
sheetname.cell(row=2, column=2).value = 'Egypt'
sheetname.cell(row=3, column=2).value = 'UpperEgypt'
sheetname.cell(row=4, column=2).value = 'Processed by JV Silva'
sheetname.cell(row=5, column=2).value = 'ICARDA'
sheetname.cell(row=6, column=2).value = 'Joao Vasco Silva, WUR'
sheetname.cell(row=9, column=1).value = 32.639637
sheetname.cell(row=9, column=2).value = 25.687243
sheetname.cell(row=9, column=3).value = -99
sheetname.cell(row=9, column=4).value = angstrom['Angst_A'].unique()[0]
sheetname.cell(row=9, column=5).value = angstrom['Angst_B'].unique()[0]
writer.book.remove_sheet(book.get_sheet_by_name('ObservedWeather'))
writer.book.save(os.path.join(input_dir, '.\wofost-inputs\weather_actual_UpperEgypt.xlsx'))
writer.book.close()

# ----------------------------------------------------------------------------------------------------------------------
# THE END
# ----------------------------------------------------------------------------------------------------------------------
